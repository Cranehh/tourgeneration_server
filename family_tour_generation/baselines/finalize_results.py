"""Step 7：整合所有 Results 章节素材到 all_metrics.xlsx + latex_tables.tex"""
import os
import pandas as pd

HERE = os.path.dirname(os.path.abspath(__file__))
RESULTS_DIR = os.path.join(HERE, "results")
EXTRA_DIR = os.path.join(RESULTS_DIR, "extra_metrics")
OUT_XLSX = os.path.join(RESULTS_DIR, "all_metrics.xlsx")
OUT_TEX = os.path.join(RESULTS_DIR, "latex_tables.tex")

DISPLAY_NAME = {
    "real": "Real",
    "main": "Ours (with Nash)",
    "main_no_nash": "Ours (w/o Nash)",
    "no_attraction": "Ours (w/o Nash, w/o Attraction)",
    "gan": "GAN",
    "wgan": "WGAN-GP",
    "vae": "VAE",
    "cgan": "CGAN",
    "cwgan": "CWGAN-GP",
    "cvae": "CVAE",
}
ORDER = ["real", "main", "main_no_nash", "no_attraction",
         "cgan", "cwgan", "cvae", "gan", "wgan", "vae"]


def order_df(df, key="model"):
    df = df.copy()
    df["_order"] = df[key].map({n: i for i, n in enumerate(ORDER)}).fillna(99)
    df = df.sort_values("_order").drop(columns="_order").reset_index(drop=True)
    df[key] = df[key].map(DISPLAY_NAME).fillna(df[key])
    return df


# ---- 加载所有结果 ----
srmse_df = pd.read_csv(os.path.join(RESULTS_DIR, "srmse_table.csv"))
fam_srmse_path = os.path.join(RESULTS_DIR, "family_srmse_table.csv")
fam_srmse_df = pd.read_csv(fam_srmse_path) if os.path.exists(fam_srmse_path) else None
coord_df = pd.read_csv(os.path.join(EXTRA_DIR, "family_coord.csv"))
trip_df = pd.read_csv(os.path.join(EXTRA_DIR, "trip_length_stats.csv"))
mode_df = pd.read_csv(os.path.join(EXTRA_DIR, "mode_share_subgroup.csv"))
spatial_path = os.path.join(RESULTS_DIR, "spatial", "zone_summary.csv")
spatial_df = pd.read_csv(spatial_path) if os.path.exists(spatial_path) else None

# 把 family SRMSE 合并进 srmse 主表
if fam_srmse_df is not None:
    srmse_df = srmse_df.merge(fam_srmse_df[["model", "family_SRMSE"]],
                                on="model", how="left")

srmse_out = order_df(srmse_df)
coord_out = order_df(coord_df)
trip_out = order_df(trip_df)

# ---- Excel 多 sheet ----
with pd.ExcelWriter(OUT_XLSX, engine="openpyxl") as w:
    srmse_out.to_excel(w, sheet_name="SRMSE主表", index=False)
    coord_out.to_excel(w, sheet_name="家庭联合性8指标", index=False)
    trip_out.to_excel(w, sheet_name="trip_length统计", index=False)
    mode_df["model"] = mode_df["model"].map(DISPLAY_NAME).fillna(mode_df["model"])
    mode_df.to_excel(w, sheet_name="mode_share子群", index=False)
    if spatial_df is not None:
        spatial_df.to_excel(w, sheet_name="目的地空间分布", index=False)

# 列宽美化
from openpyxl import load_workbook
wb = load_workbook(OUT_XLSX)
for s in wb.sheetnames:
    ws = wb[s]
    for col in ws.columns:
        max_len = max((len(str(c.value)) if c.value is not None else 0) for c in col)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 50)
wb.save(OUT_XLSX)
print(f"Saved: {OUT_XLSX}")


# ---- LaTeX 表格 ----
def fmt(x, p=3):
    if pd.isna(x):
        return "—"
    if abs(x) < 1e-4:
        return "0.000"
    return f"{x:.{p}f}"


def df_to_latex(df, caption, label, fmt_cols=None, p=3):
    cols = list(df.columns)
    fmt_cols = fmt_cols or {}
    rows = []
    rows.append(r"\begin{table}[t]")
    rows.append(r"\centering")
    rows.append(rf"\caption{{{caption}}}")
    rows.append(rf"\label{{{label}}}")
    rows.append(r"\begin{tabular}{l" + "r" * (len(cols) - 1) + "}")
    rows.append(r"\toprule")
    header = " & ".join([str(c).replace("_", r"\_") for c in cols]) + r" \\"
    rows.append(header)
    rows.append(r"\midrule")
    for _, row in df.iterrows():
        cells = []
        for c in cols:
            v = row[c]
            if isinstance(v, (int, float)) and not isinstance(v, bool):
                cells.append(fmt(v, fmt_cols.get(c, p)))
            else:
                cells.append(str(v))
        rows.append(" & ".join(cells) + r" \\")
    rows.append(r"\bottomrule")
    rows.append(r"\end{tabular}")
    rows.append(r"\end{table}")
    return "\n".join(rows)


tex = []
tex.append("% ===== Table 1: SRMSE 主表 =====")
tex.append(df_to_latex(srmse_out.drop(columns=["n_synth_valid"], errors="ignore"),
                       "SRMSE comparison across baselines (lower is better)",
                       "tab:srmse"))
tex.append("")
tex.append("% ===== Table 2: Family coordination =====")
tex.append(df_to_latex(coord_out,
                       "Family-level coordination metrics. P\\_joint = joint trip ratio; V\\_vio = vehicle constraint violation rate; H\\_vio = home-anchor violation rate; P\\_pc = parent-with-children joint trip rate.",
                       "tab:family_coord"))
tex.append("")
tex.append("% ===== Table 3: Trip length stats =====")
tex.append(df_to_latex(trip_out,
                       "Trip length statistics (standardized OD distance)",
                       "tab:trip_length"))

with open(OUT_TEX, "w", encoding="utf-8") as f:
    f.write("\n".join(tex))
print(f"Saved: {OUT_TEX}")

# ---- 终端打印汇总 ----
print()
print("=" * 70)
print("SRMSE 主表（8 模型 × 4 SRMSE 配置）")
print(srmse_out.to_string(index=False))
print()
print("=" * 70)
print("家庭联合性 4 指标")
print(coord_out.to_string(index=False))
print()
print("=" * 70)
print("Trip length 统计（标准化距离）")
print(trip_out.to_string(index=False))
