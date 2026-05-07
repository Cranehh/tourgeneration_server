"""把 srmse_table.csv 导出为 Excel（含主表 + 元数据 sheet）"""
import os
import pandas as pd

RESULTS_DIR = os.path.join(os.path.dirname(__file__), "results")
CSV_PATH = os.path.join(RESULTS_DIR, "srmse_table.csv")
XLSX_PATH = os.path.join(RESULTS_DIR, "srmse_table.xlsx")

df = pd.read_csv(CSV_PATH)

# 主表：模型中文显示名 + 类别
display_name = {
    "gan": "GAN",
    "wgan": "WGAN-GP",
    "vae": "VAE",
    "cgan": "CGAN",
    "cwgan": "CWGAN-GP",
    "cvae": "CVAE",
}
category = {
    "gan": "无条件", "wgan": "无条件", "vae": "无条件",
    "cgan": "条件", "cwgan": "条件", "cvae": "条件",
}
df_main = df.copy()
df_main.insert(1, "类别", df_main["model"].map(category))
df_main["model"] = df_main["model"].map(display_name)
df_main = df_main.rename(columns={
    "model": "模型",
    "n_synth_valid": "有效活动数",
    "SRMSE_4D": "SRMSE_4D",
    "SRMSE_5D_purpose": "SRMSE_5D_purpose",
    "SRMSE_5D_mode": "SRMSE_5D_mode",
    "SRMSE_8D_full": "SRMSE_8D_full",
})
# 在末尾加一行真实测试集参考（n_valid 真值）
df_main.loc[len(df_main)] = ["（真实测试集）", "—", 16288,
                              0.0, 0.0, 0.0, 0.0]

# 元数据 sheet
meta = pd.DataFrame([
    ["数据集", "test split (3219 个家庭，最大 8 成员 × 6 活动 = 154,512 槽位)"],
    ["真实有效活动数", "16,288 (10.54%)"],
    ["训练 epoch", 100],
    ["batch size", 256],
    ["GAN/CGAN lr", "2e-4 (Adam, betas=0.5/0.999)"],
    ["WGAN-GP/CWGAN-GP lr", "2e-4 (Adam, betas=0.5/0.9), n_critic=5, GP λ=10"],
    ["VAE/CVAE lr", "1e-3 (Adam), β=1e-4"],
    ["噪声维度 z_dim", "无条件 100, 条件 64"],
    ["MLP 隐层 hidden", 512],
    ["MLP 层数 n_layers", 3],
    ["随机种子", 42],
    ["设备", "NVIDIA RTX 4090 Laptop GPU (16GB)"],
    ["SRMSE 公式", "sqrt(mean((π_synth - π_real)^2)) / mean(π_real)，与 数据处理21.ipynb cell 108 一致"],
    ["SRMSE_4D 列", "['出发时间小时段','到达时间小时段','出发地','目的地'] 联合分布"],
    ["SRMSE_5D_purpose", "4D + ['purpose']"],
    ["SRMSE_5D_mode", "4D + ['mode']"],
    ["SRMSE_8D_full", "上述全部 + ['driver','joint'] = 8 维联合分布"],
    ["有效活动判定", "前 27 列之和 != 0（与 notebook cell 67 一致）"],
    ["主模型 SRMSE", "本表不含；请从 数据处理21.ipynb 已有结果手工填入第 7 行做对比"],
], columns=["项", "说明"])

# 写入 Excel（两个 sheet）
with pd.ExcelWriter(XLSX_PATH, engine="openpyxl") as writer:
    df_main.to_excel(writer, sheet_name="SRMSE主表", index=False)
    meta.to_excel(writer, sheet_name="实验元数据", index=False)

# 列宽美化
from openpyxl import load_workbook
wb = load_workbook(XLSX_PATH)
for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    for col in ws.columns:
        max_len = max(len(str(cell.value)) if cell.value is not None else 0 for cell in col)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 50)
wb.save(XLSX_PATH)

print(f"Excel saved: {XLSX_PATH}")
print()
print(df_main.to_string(index=False))
